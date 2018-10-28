import pickle
import openpyxl as xl
import os

def find(name, path):
	for root, dirs, files in os.walk(path):
		if name in files:
			return os.path.join(root, name)
			
def getAbsPath(file):
	current = os.getcwd()
	head, tail = os.path.split(current)
	path = find(file, head)
	return path

def loadAISC(file='shapes.xlsx'):
	path = getAbsPath(file)
	wb2 = xl.load_workbook(path)
	item2pickle = wb2.get_sheet_by_name('Database v15.0')
	return item2pickle
	
	
def database2list(file='shapes.xlsx'):
	path = getAbsPath(file)
	sheet = xl.load_workbook(path)['Database v15.0']
	
	data = []
	labels = []
	for row in sheet.iter_rows(max_row=2092):
		labels.append(row[2].value)
		row_data = []
		for cell in row:
			row_data.append(cell.value)
		
		data.append(row_data)
		
	return (data, labels)

def pickleObject(item2pickle, file='pickleditem.txt'):
	path = getAbsPath(file)
	fileObject = open(path, 'wb')
	pickle.dump(item2pickle, fileObject)
	fileObject.close()

def unPickleObject(file='pickleditem.txt'):
	path = getAbsPath(file)
	fileObject = open(path, 'rb')
	b = pickle.load(fileObject)
	return b

def main():
	a = database2list()
	pickleObject(a)

if __name__ == '__main__':
	main()