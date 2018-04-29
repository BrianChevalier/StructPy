import pickle
import openpyxl as xl

def loadAISC():
	wb2 = xl.load_workbook('shapes.xlsx')
	item2pickle = wb2.get_sheet_by_name('Database v15.0')
	return item2pickle
	
	
def database2list():
	
	sheet = xl.load_workbook('shapes.xlsx')['Database v15.0']
	
	data = []
	labels = []
	for row in sheet.iter_rows(max_row=2092):
		labels.append(row[2].value)
		row_data = []
		for cell in row:
			row_data.append(cell.value)
		
		data.append(row_data)
		
	return (data, labels)

def pickleObject(item2pickle, filename='pickleditem.txt'):
	fileObject = open(filename, 'wb')
	pickle.dump(item2pickle, fileObject)
	fileObject.close()

def unPickleObject(filename):
	fileObject = open(filename, 'rb')
	b = pickle.load(fileObject)
	return b

def main():
	a = database2list()
	pickleObject(a)

if __name__ == '__main__':
	main()
